import streamlit as st
import pandas as pd
import io
import calendar
import openpyxl
from openpyxl.styles import Font
import datetime

# 統一字型設定（與原班表一致）
CELL_FONT = Font(name="微軟正黑體", size=10)

# ==========================================
# 1. 診所排班規則與設定
# ==========================================
VALID_DOCTORS = [
    "楊忠霖", "陳逸陽", "官俊彥", "李昆晏", "劉善總", "劉庭禎", "王荷若", "鄭竣文",
    "李亞凡", "陳明志", "黃菁菁", "傅超俊", "陳昶安", "陳苡瑜", "許雅茹", "蔣宜蓁",
    "劉筑昀", "胡瑋凡"
]

# 醫師固定跟診助理（平日）
DOCTOR_ASSISTANT_MATCH = {
    "陳逸陽": "映璇",
    "王荷若": "萃屏",
    "李昆晏": "萃屏",
    "許雅茹": "和芸",
    "劉筑昀": "姿穎",
}

# 星期六特殊綁定
SATURDAY_SPECIAL_MATCH = {
    "官俊彥": "濘安",
    "陳明志": "菀庭",
}

# 助理名單
ASSISTANTS = ["映璇", "和芸", "欣寧", "萃屏", "維珍", "菀庭", "姿穎", "濘安"]

# 櫃檯優先順序名單 (只做資格篩選，排班將依診次動態平均分配)
COUNTER_PRIORITY = ["欣寧", "維珍", "和芸", "映璇", "姿穎"]

ONLY_COUNTER   = ["欣寧"]                # 只做櫃檯
NO_COUNTER     = ["萃屏", "菀庭", "濘安"]   # 不做櫃檯
NO_NIGHT_SHIFT = ["維珍"]                # 不排晚班

MAX_SHIFTS = 42  # 每位助理每月上限診次

# ==========================================
# 2. Streamlit 設定
# ==========================================
st.set_page_config(page_title="恩霖診所 - 自動排班系統", layout="wide")

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔒 恩霖診所 - 排班系統登入")
    pwd = st.text_input("請輸入密碼：", type="password")
    if st.button("登入"):
        if pwd == "115":
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("密碼錯誤！")
    st.stop()

if "timeoff_db" not in st.session_state:
    st.session_state.timeoff_db = {}

if "shift_stats" not in st.session_state:
    st.session_state.shift_stats = {}

st.title("🏥 恩霖診所 - 自動排班系統 V15 (櫃台平均分配版)")
st.info("🛡️ 已升級「動態櫃台分配」：AI 會優先將櫃台排給目前總診次最少的人，確保櫃台班平均交錯！")

tab1, tab2, tab3 = st.tabs(["📁 1. 上傳班表", "📝 2. 助理劃休", "🚀 3. AI 排班"])

# ==========================================
# Tab 1: 上傳班表
# ==========================================
with tab1:
    uploaded_file = st.file_uploader("上傳醫師班表 (Excel .xlsx)", type=["xlsx"])

    if uploaded_file:
        try:
            wb_check = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws_check = wb_check.active
            uploaded_file.seek(0)

            found_dates = []
            for row in ws_check.iter_rows():
                for cell in row:
                    if isinstance(cell.value, datetime.datetime):
                        found_dates.append(cell.value)

            if found_dates:
                ref = found_dates[0]
                st.session_state.schedule_year  = ref.year
                st.session_state.schedule_month = ref.month
                sat_days = {d.day for d in found_dates if d.weekday() == 5}
                st.session_state.saturday_dates = sat_days
                st.success(
                    f"✅ 檔案讀取完成！\n"
                    f"　偵測月份：**{ref.year} 年 {ref.month} 月**\n"
                    f"　星期六日期：{sorted(sat_days)}"
                )
            else:
                st.warning("⚠️ 班表中找不到日期資料，請確認 Excel 格式是否正確。")

        except Exception as e:
            st.error(f"讀取班表失敗：{e}")

# ==========================================
# Tab 2: 助理劃休
# ==========================================
with tab2:
    if "saturday_dates" not in st.session_state:
        st.warning("⚠️ 請先至「📁 1. 上傳班表」頁籤上傳醫師班表，系統才能自動對應正確月份與星期六。")
        st.stop()

    SATURDAY_DATES = st.session_state.saturday_dates
    year           = st.session_state.schedule_year
    month          = st.session_state.schedule_month
    days_in_month  = calendar.monthrange(year, month)[1]

    st.info(
        f"📅 目前班表月份：{year} 年 {month} 月（共 {days_in_month} 天）\n"
        f"　星期六：{sorted(SATURDAY_DATES)}"
    )

    selected_ast = st.selectbox("選擇助理：", ASSISTANTS)

    if st.session_state.get("last_schedule_month") != (year, month):
        st.session_state.timeoff_db = {}
        st.session_state.last_schedule_month = (year, month)

    if selected_ast not in st.session_state.timeoff_db:
        weekday_labels = {5: "（六）", 6: "（日）"}
        df = pd.DataFrame({
            "日期": [
                f"{i}號{weekday_labels.get(datetime.date(year, month, i).weekday(), '')}"
                for i in range(1, days_in_month + 1)
            ],
            "休整天": [False] * days_in_month,
            "早休":   [False] * days_in_month,
            "午休":   [False] * days_in_month,
            "晚休":   [False] * days_in_month,
        })
        if selected_ast in NO_NIGHT_SHIFT:
            df["晚休"] = True
        st.session_state.timeoff_db[selected_ast] = df

    saved_df = st.session_state.timeoff_db[selected_ast].copy()

    column_config = {
        "日期":   st.column_config.TextColumn("日期", width="small", disabled=True),
        "休整天": st.column_config.CheckboxColumn("✅ 休整天", width="small", help="打勾後自動勾選早、午、晚休"),
        "早休":   st.column_config.CheckboxColumn("早休", width="small"),
        "午休":   st.column_config.CheckboxColumn("午休", width="small"),
        "晚休":   st.column_config.CheckboxColumn("晚休", width="small"),
    }

    disabled_cols = []
    if selected_ast in NO_NIGHT_SHIFT:
        disabled_cols.append("晚休")

    st.caption("📌 星期六無晚班，即使勾選晚休也不會計入排班。")

    edited_df = st.data_editor(
        saved_df,
        hide_index=True,
        key=f"ed_{selected_ast}_{year}_{month}",
        column_config=column_config,
        disabled=disabled_cols,
        use_container_width=True,
    )

    for i, row in edited_df.iterrows():
        day_num = i + 1
        if row["休整天"]:
            edited_df.at[i, "早休"] = True
            edited_df.at[i, "午休"] = True
            if day_num not in SATURDAY_DATES and selected_ast not in NO_NIGHT_SHIFT:
                edited_df.at[i, "晚休"] = True
        if day_num in SATURDAY_DATES:
            edited_df.at[i, "晚休"] = False

    col_btn, col_status = st.columns([1, 3])
    with col_btn:
        save_clicked = st.button(f"💾 儲存 {selected_ast} 休假", type="primary")

    if save_clicked:
        st.session_state.timeoff_db[selected_ast] = edited_df.copy()
        with col_status:
            total_early = edited_df["早休"].sum()
            total_noon  = edited_df["午休"].sum()
            total_night = edited_df["晚休"].sum()
            full_days   = edited_df["休整天"].sum()
            st.success(
                f"✅ {selected_ast} 休假已儲存！　"
                f"休整天：{full_days} 天　早休：{total_early} 次　"
                f"午休：{total_noon} 次　晚休：{total_night} 次"
            )

    st.divider()
    st.caption("📌 目前已儲存的休假紀錄")
    saved = st.session_state.timeoff_db[selected_ast]
    full_day_list = [f"{i+1}號" for i, r in saved.iterrows() if r["休整天"]]
    early_list    = [f"{i+1}號" for i, r in saved.iterrows() if r["早休"] and not r["休整天"]]
    noon_list     = [f"{i+1}號" for i, r in saved.iterrows() if r["午休"] and not r["休整天"]]
    night_list    = [f"{i+1}號" for i, r in saved.iterrows() if r["晚休"] and not r["休整天"]]

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown("**🔴 休整天**")
        st.write("、".join(full_day_list) if full_day_list else "（無）")
    with c2:
        st.markdown("**🟠 僅早休**")
        st.write("、".join(early_list) if early_list else "（無）")
    with c3:
        st.markdown("**🟡 僅午休**")
        st.write("、".join(noon_list) if noon_list else "（無）")
    with c4:
        st.markdown("**🟤 僅晚休**")
        st.write("、".join(night_list) if night_list else "（無）")

# ==========================================
# Tab 3: 排班
# ==========================================
with tab3:

    if st.session_state.shift_stats:
        st.subheader("📊 上次排班診次統計")
        stats = st.session_state.shift_stats
        cols = st.columns(len(ASSISTANTS))
        for idx, ast in enumerate(ASSISTANTS):
            count = stats.get(ast, 0)
            over  = count > MAX_SHIFTS
            cols[idx].metric(
                label=ast,
                value=f"{count} 診",
                delta=f"{'⚠️ 超額 ' + str(count - MAX_SHIFTS) + ' 診' if over else '正常'}",
                delta_color="inverse" if over else "normal",
            )
        st.divider()

    if st.button("🚀 開始智慧排班", type="primary"):
        if not uploaded_file:
            st.error("請先上傳班表！")
        elif "saturday_dates" not in st.session_state:
            st.error("請先至「📁 1. 上傳班表」上傳班表，讓系統解析月份資訊。")
        else:
            with st.spinner("正在解析班表並排班中..."):
                try:
                    uploaded_file.seek(0)
                    wb = openpyxl.load_workbook(uploaded_file)
                    ws = wb.active

                    rotation_counter = {a: 0 for a in ASSISTANTS}

                    # ── 輔助函式 ──────────────────────────────────────
                    def is_on_leave(ast_name, day_num, shift_type):
                        if ast_name not in st.session_state.timeoff_db: return False
                        if not (1 <= day_num <= 31): return False
                        db  = st.session_state.timeoff_db[ast_name]
                        col = "早休" if "早" in shift_type else ("午休" if "午" in shift_type else "晚休")
                        try:
                            return bool(db[col].iloc[day_num - 1])
                        except Exception: return False

                    def over_limit(ast_name):
                        return rotation_counter[ast_name] >= MAX_SHIFTS

                    def write_cell(row, col, value):
                        cell = ws.cell(row, col)
                        cell.value = value
                        cell.font  = CELL_FONT

                    def find_doctors_in_cell(cell_value):
                        if not cell_value: return []
                        text = str(cell_value).replace(" ", "")
                        if "休" in text: return []
                        return [doc for doc in VALID_DOCTORS if doc in text]

                    def ok_for_assist(candidate, working_now, day_num, shift_name, daily_counts):
                        return (
                            candidate not in working_now
                            and not is_on_leave(candidate, day_num, shift_name)
                            and not over_limit(candidate)
                            and daily_counts.get(candidate, 0) < 2
                            and candidate not in ONLY_COUNTER
                            and candidate not in NO_COUNTER
                            and not ("晚" in shift_name and candidate in NO_NIGHT_SHIFT)
                        )

                    def ok_for_counter(candidate, working_now, day_num, shift_name, daily_counts):
                        return (
                            candidate not in working_now
                            and not is_on_leave(candidate, day_num, shift_name)
                            and not over_limit(candidate)
                            and daily_counts.get(candidate, 0) < 2
                            and not ("晚" in shift_name and candidate in NO_NIGHT_SHIFT)
                        )

                    # ── 找出所有週塊起始行 ────────────────────────────
                    enlin_rows = []
                    for r in range(1, ws.max_row + 1):
                        v = ws.cell(r, 1).value
                        if v and "恩霖" in str(v):
                            enlin_rows.append(r)

                    if not enlin_rows:
                        st.error("找不到班表週塊，請確認上傳的檔案格式正確。")
                        st.stop()

                    # ── 處理每個週塊 ──────────────────────────────────
                    for week_idx, week_start in enumerate(enlin_rows):
                        week_end = (enlin_rows[week_idx + 1] if week_idx + 1 < len(enlin_rows) else ws.max_row + 1)

                        date_cols = []
                        for c in range(2, ws.max_column + 1, 2):
                            v = ws.cell(week_start, c).value
                            if isinstance(v, datetime.datetime):
                                date_cols.append((c, c + 1, v.day, v.weekday() == 5))

                        if not date_cols: continue

                        row_labels = {}
                        for r in range(week_start, week_end):
                            v1 = str(ws.cell(r, 1).value or "").replace(" ", "")
                            if   "早班"  in v1: row_labels.setdefault("早班",  r)
                            elif "早櫃2" in v1: row_labels.setdefault("早櫃2", r)
                            elif "早櫃1" in v1: row_labels.setdefault("早櫃1", r)
                            elif "午班"  in v1: row_labels.setdefault("午班",  r)
                            elif "午櫃2" in v1: row_labels.setdefault("午櫃2", r)
                            elif "午櫃1" in v1: row_labels.setdefault("午櫃1", r)
                            elif "晚班"  in v1: row_labels.setdefault("晚班",  r)
                            elif "晚櫃2" in v1: row_labels.setdefault("晚櫃2", r)
                            elif "晚櫃1" in v1: row_labels.setdefault("晚櫃1", r)
                            elif "備註"  in v1: row_labels.setdefault("備註",  r); break

                        shifts = []
                        if "早班" in row_labels and "早櫃2" in row_labels:
                            doc_rows = [r for r in range(row_labels["早班"] + 1, row_labels["早櫃2"]) if str(ws.cell(r, 1).value or "").strip() in ("11", "21", "22", "23", "24")]
                            counter_rows = [x for x in [row_labels.get("早櫃2"), row_labels.get("早櫃1")] if x]
                            shifts.append(("早班", doc_rows, counter_rows))

                        if "早櫃1" in row_labels and "午櫃2" in row_labels:
                            doc_rows = [r for r in range(row_labels["早櫃1"] + 1, row_labels["午櫃2"]) if str(ws.cell(r, 1).value or "").strip() in ("11", "21", "22", "23", "24")]
                            counter_rows = [x for x in [row_labels.get("午櫃2"), row_labels.get("午櫃1")] if x]
                            shifts.append(("午班", doc_rows, counter_rows))

                        if "晚班" in row_labels and "晚櫃2" in row_labels:
                            doc_rows = [r for r in range(row_labels["晚班"] + 1, row_labels["晚櫃2"]) if str(ws.cell(r, 1).value or "").strip() in ("11", "21", "22", "23", "24")]
                            counter_rows = [x for x in [row_labels.get("晚櫃2"), row_labels.get("晚櫃1")] if x]
                            shifts.append(("晚班", doc_rows, counter_rows))

                        # 3. 以「天」為單位排班
                        for doc_col, asst_col, day_num, is_saturday in date_cols:
                            doc_day_memory = {}
                            daily_shift_counts = {a: 0 for a in ASSISTANTS}

                            for shift_name, doc_rows, counter_rows in shifts:
                                docs_this_shift = []
                                for r in doc_rows:
                                    for doc in find_doctors_in_cell(ws.cell(r, doc_col).value):
                                        docs_this_shift.append((r, doc))

                                if not docs_this_shift: continue

                                working_now = set()

                                # A. 預排保護機制 (掃描本班次已填寫的助理)
                                for r in doc_rows + counter_rows:
                                    ea = ws.cell(r, asst_col).value
                                    if ea and str(ea).strip() in ASSISTANTS:
                                        ast = str(ea).strip()
                                        working_now.add(ast)
                                        daily_shift_counts[ast] += 1
                                        rotation_counter[ast] += 1
                                        
                                        if r in doc_rows:
                                            docs_here = find_doctors_in_cell(ws.cell(r, doc_col).value)
                                            for doc in docs_here:
                                                doc_day_memory[doc] = ast

                                # B. 跟診助理分配
                                for r, doc in docs_this_shift:
                                    ea = ws.cell(r, asst_col).value
                                    if ea: continue 

                                    assigned = None

                                    if doc in doc_day_memory:
                                        cand = doc_day_memory[doc]
                                        if ok_for_assist(cand, working_now, day_num, shift_name, daily_shift_counts):
                                            assigned = cand
                                    if not assigned and is_saturday and doc in SATURDAY_SPECIAL_MATCH:
                                        cand = SATURDAY_SPECIAL_MATCH[doc]
                                        if ok_for_assist(cand, working_now, day_num, shift_name, daily_shift_counts):
                                            assigned = cand
                                    if not assigned and not is_saturday and doc in DOCTOR_ASSISTANT_MATCH:
                                        cand = DOCTOR_ASSISTANT_MATCH[doc]
                                        if ok_for_assist(cand, working_now, day_num, shift_name, daily_shift_counts):
                                            assigned = cand
                                    if not assigned:
                                        pool = sorted(
                                            [a for a in ASSISTANTS if ok_for_assist(a, working_now, day_num, shift_name, daily_shift_counts)],
                                            key=lambda a: rotation_counter[a]
                                        )
                                        if pool: assigned = pool[0]

                                    if assigned:
                                        write_cell(r, asst_col, assigned)
                                        working_now.add(assigned)
                                        doc_day_memory[doc] = assigned
                                        rotation_counter[assigned] += 1
                                        daily_shift_counts[assigned] += 1
                                    else:
                                        write_cell(r, asst_col, "缺")

                                # 🌟 C. 櫃檯分配 (核心優化：動態平均分配) 🌟
                                need_two = len(docs_this_shift) >= 4
                                target_count = 2 if need_two else 1
                                
                                pre_assigned_counters = sum(
                                    1 for cr in counter_rows 
                                    if ws.cell(cr, asst_col).value and str(ws.cell(cr, asst_col).value).strip() in ASSISTANTS
                                )
                                
                                counters_to_add = target_count - pre_assigned_counters
                                assigned_new_counters = []

                                if counters_to_add > 0:
                                    # 1. 抓出可以站櫃檯的優先名單，並「依照目前總診次由少到多」動態排序
                                    available_counters = [
                                        cand for cand in COUNTER_PRIORITY
                                        if ok_for_counter(cand, working_now, day_num, shift_name, daily_shift_counts)
                                    ]
                                    available_counters.sort(key=lambda a: rotation_counter.get(a, 0))

                                    # 2. 優先給目前時數最少的人
                                    for cand in available_counters:
                                        if len(assigned_new_counters) >= counters_to_add: break
                                        assigned_new_counters.append(cand)
                                        working_now.add(cand)
                                        rotation_counter[cand] += 1
                                        daily_shift_counts[cand] += 1

                                    # 3. 如果優先名單用完了還是缺人，從一般池子裡找
                                    while len(assigned_new_counters) < counters_to_add:
                                        pool = sorted(
                                            [a for a in ASSISTANTS if ok_for_counter(a, working_now, day_num, shift_name, daily_shift_counts) and a not in NO_COUNTER],
                                            key=lambda a: rotation_counter[a]
                                        )
                                        if pool:
                                            assigned_new_counters.append(pool[0])
                                            working_now.add(pool[0])
                                            rotation_counter[pool[0]] += 1
                                            daily_shift_counts[pool[0]] += 1
                                        else:
                                            assigned_new_counters.append("缺")
                                            break

                                # 寫入未被手動預排的櫃檯格子
                                c_idx = 0
                                filled_count = pre_assigned_counters
                                for cr in counter_rows:
                                    ea = ws.cell(cr, asst_col).value
                                    if not ea or str(ea).strip() not in ASSISTANTS:
                                        if c_idx < len(assigned_new_counters):
                                            write_cell(cr, asst_col, assigned_new_counters[c_idx])
                                            c_idx += 1
                                            filled_count += 1
                                        elif filled_count < target_count:
                                            write_cell(cr, asst_col, "缺")
                                            filled_count += 1

                    # ── 儲存與顯示診次統計 ──────────────────────────────────
                    st.session_state.shift_stats = dict(rotation_counter)

                    st.subheader("📊 本次排班診次統計")
                    st.caption(f"每位助理上限：{MAX_SHIFTS} 診／月　｜　已達上限者自動標「缺」")

                    cols = st.columns(len(ASSISTANTS))
                    for idx, ast in enumerate(ASSISTANTS):
                        count = rotation_counter.get(ast, 0)
                        over  = count > MAX_SHIFTS
                        cols[idx].metric(
                            label=ast,
                            value=f"{count} 診",
                            delta=(f"⚠️ 超額 {count - MAX_SHIFTS} 診" if over else f"剩餘 {MAX_SHIFTS - count} 診"),
                            delta_color="inverse" if over else "normal",
                        )

                    st.subheader("📈 診次進度")
                    bar_cols = st.columns(len(ASSISTANTS))
                    for idx, ast in enumerate(ASSISTANTS):
                        count = rotation_counter.get(ast, 0)
                        pct   = min(count / MAX_SHIFTS, 1.0)
                        with bar_cols[idx]:
                            st.write(f"**{ast}**")
                            st.progress(pct, text=f"{count}/{MAX_SHIFTS}")

                    st.divider()

                    # ── 輸出 Excel ────────────────────────────────────
                    s_year  = st.session_state.schedule_year
                    s_month = st.session_state.schedule_month
                    output  = io.BytesIO()
                    wb.save(output)
                    st.success("✅ 排班完成！櫃台班次已成功平均分配。")
                    st.download_button(
                        "📥 下載排班結果",
                        output.getvalue(),
                        f"恩霖診所_{s_year}年{s_month}月排班結果.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                except Exception as e:
                    st.error(f"排班過程發生錯誤：{e}")
                    import traceback
                    st.code(traceback.format_exc())
